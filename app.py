from flask import Flask, render_template, jsonify, send_from_directory, request, redirect, url_for, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_migrate import Migrate
from werkzeug.security import check_password_hash, generate_password_hash
from flask_wtf.csrf import CSRFProtect, generate_csrf
import json
import pandas as pd
import os
from models import db, TestResult
from datetime import datetime, date, timedelta
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io

app = Flask(__name__)

# Настройка конфигурации
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')
app.config['ADMIN_USERNAME'] = os.environ.get('ADMIN_USERNAME')
app.config['ADMIN_PASSWORD'] = os.environ.get('ADMIN_PASSWORD')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///test_results.db')
if app.config['SQLALCHEMY_DATABASE_URI'].startswith("postgres://"):
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Инициализация CSRF защиты
csrf = CSRFProtect(app)

# Настройка CSRF для AJAX запросов
@app.after_request
def add_csrf_header(response):
    if 'text/html' in response.headers['Content-Type']:
        response.set_cookie('csrf_token', generate_csrf())
    return response

# Добавляем фильтр для преобразования datetime в date
@app.template_filter('datetime_to_date')
def datetime_to_date(value):
    return value.date() if value else None

# Инициализация расширений
db.init_app(app)
migrate = Migrate(app, db)

# Инициализация Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'admin_login'

# Создаем таблицы при первом запуске, если их нет
with app.app_context():
    db.create_all()

def load_questions():
    try:
        with open('test-questions.json', 'r', encoding='utf-8') as file:
            data = json.load(file)
            # Получаем все вопросы из всех категорий
            questions = []
            for category_key, category_data in data['categories'].items():
                category_questions = category_data['questions']
                for question in category_questions:
                    questions.append({
                        'text': question['text'],
                        'id': question['id'],
                        'category': {
                            'key': category_key,
                            'name': category_data['name'],
                            'description': category_data['description']
                        }
                    })
            return questions
    except Exception as e:
        print(f"Ошибка при чтении файла test-questions.json: {e}")
        return []

@app.route('/')
def index():
    questions = load_questions()
    return render_template('index.html', question=questions[0] if questions else None)

@app.route('/api/questions')
def get_questions():
    questions = load_questions()
    return jsonify(questions)

@app.route('/results')
def results():
    try:
        # Получаем данные регистрации и результаты теста из сессии
        registration_data = session.get('registration_data', {})
        test_results = session.get('test_results', {})
        
        if not test_results:
            return redirect(url_for('register'))
            
        # Получаем описание типа из CSV файла
        try:
            df = pd.read_csv('karta_talantov_type.csv')
            type_data = df.to_dict('records')
            type_description = None
            
            if test_results.get('type_code'):
                type_match = next((t for t in type_data if test_results['type_code'] in t['type_simbol'].split('/')), None)
                if type_match:
                    type_description = format_description(type_match['discription'])
        except Exception as e:
            print(f"Ошибка при чтении описания типа: {str(e)}")
            type_description = None
        
        # После успешного отображения результатов очищаем данные сессии
        session.pop('registration_data', None)
        session.pop('test_results', None)
        
        return render_template('results.html',
                             registration_data=registration_data,
                             test_results=test_results,
                             type_description=type_description)
    except Exception as e:
        print(f"Ошибка при отображении результатов: {e}")
        return "Ошибка при отображении страницы результатов", 500

@app.route('/api/type-descriptions')
def get_type_descriptions():
    try:
        # Читаем CSV файл
        df = pd.read_csv('karta_talantov_type.csv')
        # Преобразуем в список словарей
        type_descriptions = df.to_dict('records')
        return jsonify(type_descriptions)
    except Exception as e:
        print(f"Ошибка при чтении CSV файла: {e}")
        return jsonify([])

# Добавим обработку статических файлов
@app.route('/<path:filename>')
def serve_static(filename):
    return send_from_directory(app.static_folder, filename)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        # Сохраняем данные формы в сессии
        session['registration_data'] = {
            'parent_name': request.form.get('parent_name'),
            'child_name': request.form.get('child_name'),
            'phone': request.form.get('phone'),
            'grade': int(request.form.get('grade')),  # Преобразуем в число
            'city': request.form.get('city')  # Добавляем город
        }
        return redirect(url_for('index'))
    return render_template('register.html')

@app.route('/api/save-results', methods=['POST'])
def save_results():
    try:
        print("Получен запрос на сохранение результатов")
        data = request.get_json()
        print("Полученные данные:", data)
        
        registration_data = session.get('registration_data', {})
        print("Данные регистрации из сессии:", registration_data)
        
        # Проверяем наличие всех необходимых данных
        if not registration_data:
            print("Ошибка: нет данных регистрации")
            return jsonify({'success': False, 'error': 'Нет данных регистрации'}), 400
            
        required_fields = ['parent_name', 'child_name', 'phone', 'grade', 'city']
        missing_fields = [field for field in required_fields if not registration_data.get(field)]
        if missing_fields:
            print("Ошибка: отсутствуют обязательные поля:", missing_fields)
            return jsonify({
                'success': False, 
                'error': f'Отсутствуют обязательные поля: {", ".join(missing_fields)}'
            }), 400

        if not data.get('personality_type') or not data.get('type_code'):
            print("Ошибка: отсутствуют данные о типе личности")
            return jsonify({
                'success': False,
                'error': 'Отсутствуют данные о типе личности'
            }), 400
        
        print("Создаем запись в базе данных")
        # Создаем новую запись в базе данных
        test_result = TestResult(
            parent_name=registration_data.get('parent_name'),
            child_name=registration_data.get('child_name'),
            phone=registration_data.get('phone'),
            grade=registration_data.get('grade'),
            city=registration_data.get('city'),
            personality_type=data.get('personality_type'),
            type_code=data.get('type_code'),
            is_completed=True
        )
        
        try:
            print("Сохраняем в базу данных")
            db.session.add(test_result)
            db.session.commit()
            
            print("Сохраняем результаты в сессии")
            # Сохраняем результаты в сессии для отображения на странице результатов
            session['test_results'] = {
                'personality_type': data.get('personality_type'),
                'type_code': data.get('type_code'),
                'scores': data.get('scores', {})
            }
            
            print("Успешно сохранено")
            return jsonify({'success': True})
        except Exception as db_error:
            db.session.rollback()
            print(f"Ошибка базы данных: {str(db_error)}")
            return jsonify({
                'success': False,
                'error': 'Ошибка при сохранении в базу данных'
            }), 500
        
    except Exception as e:
        print(f"Общая ошибка: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Простая модель пользователя для админа
class Admin(UserMixin):
    def __init__(self, id):
        self.id = id
        self.name = app.config['ADMIN_USERNAME']

# Создаем единственного админа
admin = Admin(1)

@login_manager.user_loader
def load_user(user_id):
    if int(user_id) == 1:
        return admin
    return None

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == app.config['ADMIN_USERNAME'] and password == app.config['ADMIN_PASSWORD']:
            login_user(admin)
            return redirect(url_for('admin_dashboard'))
        return render_template('admin_login.html', error='Неверный логин или пароль')
    return render_template('admin_login.html')

@app.route('/admin/logout')
@login_required
def admin_logout():
    logout_user()
    return redirect(url_for('admin_login'))

@app.route('/admin')
@login_required
def admin_redirect():
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    try:
        # Получаем параметры фильтров из запроса
        city = request.args.get('city', '')
        grade = request.args.get('grade', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')

        # Базовый запрос
        query = TestResult.query

        # Применяем фильтры
        if city:
            query = query.filter(TestResult.city == city)
        if grade:
            query = query.filter(TestResult.grade == int(grade))
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(TestResult.test_date >= date_from)
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            date_to = date_to + timedelta(days=1)  # Включаем весь день
            query = query.filter(TestResult.test_date < date_to)

        # Получаем отфильтрованные результаты
        results = query.order_by(TestResult.test_date.desc()).all()

        # Получаем уникальные города и классы для фильтров
        cities = db.session.query(TestResult.city).distinct().order_by(TestResult.city).all()
        cities = [city[0] for city in cities]
        
        grades = db.session.query(TestResult.grade).distinct().order_by(TestResult.grade).all()
        grades = [grade[0] for grade in grades]

        # Подсчитываем количество тестов за сегодня
        today = datetime.now().date()
        tests_today = TestResult.query.filter(
            db.func.date(TestResult.test_date) == today
        ).count()

        selected_filters = {
            'city': city,
            'grade': grade,
            'date_from': date_from,
            'date_to': date_to
        }

        return render_template('admin_dashboard.html',
                             results=results,
                             tests_today=tests_today,
                             cities=cities,
                             grades=grades,
                             selected_filters=selected_filters)
    except Exception as e:
        print(f"Ошибка в админ-панели: {e}")
        return "Ошибка при загрузке админ-панели", 500

def format_description(description):
    sections = {
        'Описание типа личности': [],
        'Характеристика': [],
        'Интересы': [],
        'Рекомендации для родителей': [],
        'Ресурсы и материалы': [],
        'Какая школа подходит?': [],
        'Возможные профессии в будущем': [],
        'Советы по воспитанию и развитию': [],
        'Итог для родителей': []
    }

    current_section = 'Описание типа личности'
    lines = description.split('\n')

    for line in lines:
        trimmed_line = line.strip()
        # Убираем черную точку в начале строки, если она есть
        clean_line = re.sub(r'^[•·]?\s*', '', trimmed_line)
        
        if trimmed_line in sections:
            current_section = trimmed_line
        elif clean_line:
            sections[current_section].append(clean_line)

    formatted_html = []
    for title, content in sections.items():
        if content:
            section_html = f'<section><h5>{title}</h5><div class="section-content">'
            for line in content:
                # Проверяем, является ли строка элементом списка
                if line.startswith('•') or re.match(r'^[-•·]\s', line):
                    line_content = re.sub(r'^[-•·]\s', '', line)
                    section_html += f'<li>{line_content}</li>'
                elif ':' in line and not line.endswith(':'):
                    label, text = line.split(':', 1)
                    section_html += f'<div class="highlight-box"><strong>{label}:</strong>{text}</div>'
                else:
                    section_html += f'<p>{line}</p>'
            section_html += '</div></section>'
            formatted_html.append(section_html)

    return '\n'.join(formatted_html)

@app.route('/admin/result/<int:result_id>')
@login_required
def view_result(result_id):
    # Получаем результат теста по ID
    result = TestResult.query.get_or_404(result_id)
    
    # Получаем описание типа из CSV файла
    try:
        df = pd.read_csv('karta_talantov_type.csv')
        type_data = df.to_dict('records')
        
        # Находим соответствующее описание
        type_description = None
        if result.type_code:
            type_match = next((t for t in type_data if result.type_code in t['type_simbol'].split('/')), None)
            if type_match:
                type_description = format_description(type_match['discription'])
    except Exception as e:
        print(f"Ошибка при чтении описания типа: {str(e)}")
        type_description = None
    
    return render_template('view_result.html', 
                         result=result,
                         type_description=type_description)

@app.route('/admin/export-excel')
@login_required
def export_excel():
    try:
        # Получаем те же параметры фильтров, что и в админ-панели
        city = request.args.get('city', '')
        grade = request.args.get('grade', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')

        # Базовый запрос
        query = TestResult.query

        # Применяем фильтры
        if city:
            query = query.filter(TestResult.city == city)
        if grade:
            query = query.filter(TestResult.grade == int(grade))
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(TestResult.test_date >= date_from)
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            date_to = date_to + timedelta(days=1)
            query = query.filter(TestResult.test_date < date_to)

        # Получаем отфильтрованные результаты
        results = query.order_by(TestResult.test_date.desc()).all()

        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты тестирования"

        # Заголовки
        headers = [
            'Дата', 'Родитель', 'Ребенок', 'Телефон', 'Класс',
            'Город/Область', 'Статус', 'Тип личности', 'Код'
        ]

        # Стили для заголовков
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Записываем данные
        for row, result in enumerate(results, 2):
            ws.cell(row=row, column=1, value=result.test_date.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row, column=2, value=result.parent_name)
            ws.cell(row=row, column=3, value=result.child_name)
            ws.cell(row=row, column=4, value=result.phone)
            ws.cell(row=row, column=5, value=f"{result.grade} класс")
            ws.cell(row=row, column=6, value=result.city)
            ws.cell(row=row, column=7, value="Завершен" if result.is_completed else "Не завершен")
            ws.cell(row=row, column=8, value=result.personality_type or "—")
            ws.cell(row=row, column=9, value=result.type_code or "—")

        # Автоматическая ширина столбцов
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Сохраняем в буфер
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'результаты_тестирования_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )

    except Exception as e:
        print(f"Ошибка при экспорте в Excel: {e}")
        return "Ошибка при экспорте данных", 500

if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5001) 